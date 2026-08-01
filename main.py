from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


BASE_DIR = Path(__file__).resolve().parent
INVENTORY_FILE = BASE_DIR / "Inventory.xlsx"
ORDERS_FILE = BASE_DIR / "Orders.xlsx"
OUTPUT_FILE = BASE_DIR / "UpdatedInventory.xlsx"
REPORT_FILE = BASE_DIR / "InventoryReport.pdf"
LOG_FILE = BASE_DIR / "inventory_sync.log"


def configure_logging() -> None:
    logging.basicConfig(
        filename=LOG_FILE,
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
    )


def load_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    if not INVENTORY_FILE.exists():
        raise FileNotFoundError(f"Missing file: {INVENTORY_FILE.name}")
    if not ORDERS_FILE.exists():
        raise FileNotFoundError(f"Missing file: {ORDERS_FILE.name}")

    inventory = pd.read_excel(INVENTORY_FILE)
    orders = pd.read_excel(ORDERS_FILE)

    required_inventory = {
        "Product ID",
        "Product Name",
        "Category",
        "Current Stock",
        "Minimum Stock",
        "Unit Price",
    }
    required_orders = {"Order ID", "Product ID", "Quantity Sold"}

    missing_inventory = required_inventory - set(inventory.columns)
    missing_orders = required_orders - set(orders.columns)

    if missing_inventory:
        raise ValueError(f"Inventory is missing columns: {sorted(missing_inventory)}")
    if missing_orders:
        raise ValueError(f"Orders is missing columns: {sorted(missing_orders)}")

    return inventory, orders


def update_inventory(
    inventory: pd.DataFrame, orders: pd.DataFrame
) -> tuple[pd.DataFrame, list[str]]:
    sales_by_product = (
        orders.groupby("Product ID", as_index=False)["Quantity Sold"].sum()
    )

    updated = inventory.merge(
        sales_by_product,
        on="Product ID",
        how="left",
    )
    updated["Quantity Sold"] = updated["Quantity Sold"].fillna(0).astype(int)

    warnings: list[str] = []
    updated["Previous Stock"] = updated["Current Stock"]
    updated["Current Stock"] = (
        updated["Current Stock"] - updated["Quantity Sold"]
    )

    negative_mask = updated["Current Stock"] < 0
    if negative_mask.any():
        for _, row in updated.loc[negative_mask].iterrows():
            warnings.append(
                f"{row['Product ID']} ({row['Product Name']}): "
                f"sales exceeded available stock."
            )
        updated.loc[negative_mask, "Current Stock"] = 0

    updated["Status"] = updated.apply(
        lambda row: (
            "Out of Stock"
            if row["Current Stock"] == 0
            else "Low Stock"
            if row["Current Stock"] <= row["Minimum Stock"]
            else "Available"
        ),
        axis=1,
    )

    columns = [
        "Product ID",
        "Product Name",
        "Category",
        "Previous Stock",
        "Quantity Sold",
        "Current Stock",
        "Minimum Stock",
        "Unit Price",
        "Status",
    ]
    return updated[columns], warnings


def save_excel(updated: pd.DataFrame) -> None:
    updated.to_excel(OUTPUT_FILE, index=False, sheet_name="Updated Inventory")

    workbook = load_workbook(OUTPUT_FILE)
    sheet = workbook["Updated Inventory"]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    status_column = None
    for cell in sheet[1]:
        if cell.value == "Status":
            status_column = cell.column
            break

    status_colors = {
        "Available": ("D9EAD3", "274E13"),
        "Low Stock": ("FFF2CC", "7F6000"),
        "Out of Stock": ("F4CCCC", "9C0006"),
    }

    if status_column:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=status_column)
            fill_color, font_color = status_colors.get(
                cell.value, ("FFFFFF", "000000")
            )
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(color=font_color, bold=cell.value != "Available")

    widths = {
        "A": 14,
        "B": 30,
        "C": 18,
        "D": 16,
        "E": 15,
        "F": 15,
        "G": 15,
        "H": 14,
        "I": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    workbook.save(OUTPUT_FILE)


def generate_pdf(updated: pd.DataFrame, warnings: list[str]) -> None:
    document = SimpleDocTemplate(
        str(REPORT_FILE),
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Inventory Sync Automation Report", styles["Title"]),
        Spacer(1, 12),
        Paragraph(
            f"Products processed: {len(updated)}",
            styles["BodyText"],
        ),
        Paragraph(
            f"Units sold: {int(updated['Quantity Sold'].sum())}",
            styles["BodyText"],
        ),
        Paragraph(
            f"Low-stock products: {(updated['Status'] == 'Low Stock').sum()}",
            styles["BodyText"],
        ),
        Paragraph(
            f"Out-of-stock products: {(updated['Status'] == 'Out of Stock').sum()}",
            styles["BodyText"],
        ),
        Spacer(1, 16),
    ]

    table_data = [
        [
            "Product",
            "Previous",
            "Sold",
            "Current",
            "Status",
        ]
    ]

    for _, row in updated.iterrows():
        table_data.append(
            [
                str(row["Product Name"]),
                str(row["Previous Stock"]),
                str(row["Quantity Sold"]),
                str(row["Current Stock"]),
                str(row["Status"]),
            ]
        )

    table = Table(table_data, repeatRows=1, colWidths=[190, 65, 50, 60, 85])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 1), (-2, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                    colors.white,
                    colors.HexColor("#F3F6F9"),
                ]),
            ]
        )
    )
    elements.append(table)

    if warnings:
        elements.extend(
            [
                Spacer(1, 16),
                Paragraph("Warnings", styles["Heading2"]),
            ]
        )
        for warning in warnings:
            elements.append(Paragraph(f"• {warning}", styles["BodyText"]))

    document.build(elements)


def main() -> None:
    configure_logging()
    logging.info("Inventory synchronization started.")

    try:
        inventory, orders = load_data()
        updated, warnings = update_inventory(inventory, orders)
        save_excel(updated)
        generate_pdf(updated, warnings)

        logging.info("Inventory synchronization completed successfully.")
        print("Inventory synchronization completed.")
        print(f"Created: {OUTPUT_FILE.name}")
        print(f"Created: {REPORT_FILE.name}")

        if warnings:
            print("\nWarnings:")
            for warning in warnings:
                print(f"- {warning}")

    except Exception as exc:
        logging.exception("Inventory synchronization failed.")
        print(f"Error: {exc}")
        raise


if __name__ == "__main__":
    main()
